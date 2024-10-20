import tkinter as tk
from tkinter import filedialog, messagebox
import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
import code128
import io
import os

class BarcodeApp:
    def __init__(self, master):
        self.master = master
        master.title("Barcode Management App")
        
        # Set window size
        master.geometry("400x200")
        
        # Set window background color
        master.configure(bg="white")

        # Create buttons
        self.create_barcode_button = tk.Button(master, text="Create a Barcode", command=self.open_create_barcode_window, 
                                                bg="#ff9800", fg="black", font=("Arial", 12, "bold"))
        self.create_barcode_button.pack(pady=20)

        self.scan_barcode_button = tk.Button(master, text="Scan a Barcode", command=self.scan_barcode, 
                                              bg="#ff9800", fg="black", font=("Arial", 12, "bold"))
        self.scan_barcode_button.pack(pady=20)

    def open_create_barcode_window(self):
        self.create_window = tk.Toplevel(self.master)
        self.create_window.title("Create Barcodes")
        self.create_window.geometry("500x500")

        self.input_frame = tk.Frame(self.create_window)
        self.input_frame.pack(pady=20)

        self.entries = []  # List to hold input entries
        
        # Add initial entry fields
        self.add_entry()

        # Add buttons
        self.add_entry_button = tk.Button(self.create_window, text="Add Item", command=self.add_entry)
        self.add_entry_button.pack(pady=5)

        self.save_barcode_button = tk.Button(self.create_window, text="Save Barcodes in Excel", command=self.save_barcodes)
        self.save_barcode_button.pack(pady=5)

    def add_entry(self):
        entry_frame = tk.Frame(self.input_frame)
        entry_frame.pack(pady=5)

        item_code_entry = self.create_placeholder_entry(entry_frame, "Item Code")
        description_entry = self.create_placeholder_entry(entry_frame, "Description")

        self.entries.append((item_code_entry, description_entry))

    def create_placeholder_entry(self, parent, placeholder):
        entry = tk.Entry(parent, width=15)
        entry.pack(side=tk.LEFT, padx=5)

        # Set the placeholder text
        entry.insert(0, placeholder)
        entry.bind("<FocusIn>", lambda event: self.clear_placeholder(event, placeholder))
        entry.bind("<FocusOut>", lambda event: self.set_placeholder(event, placeholder))
        
        return entry

    def clear_placeholder(self, event, placeholder):
        if event.widget.get() == placeholder:
            event.widget.delete(0, tk.END)  # Clear the entry field

    def set_placeholder(self, event, placeholder):
        if event.widget.get() == '':
            event.widget.insert(0, placeholder)  # Set the placeholder back if empty

    def save_barcodes(self):
        # Ask for the directory to save Excel
        save_dir = filedialog.askdirectory(title="Select Directory to Save Excel")
        if not save_dir:
            messagebox.showwarning("No Directory Selected", "Please select a directory to save the Excel file.")
            return

        excel_file_path = os.path.join(save_dir, "database_with_barcodes.xlsx")

        # Create a new workbook and select the active worksheet
        wb = Workbook()
        ws = wb.active
        ws.append(['Item Code', 'Description', 'Barcode'])

        # Generate barcodes for each entry
        for item_code_entry, description_entry in self.entries:
            item_code = item_code_entry.get().strip()
            description = description_entry.get().strip()
            if item_code and item_code != "Item Code":  # Avoid saving placeholder text
                # Add item code and description to Excel
                ws.append([item_code, description])

                # Generate barcode image using the code128 library
                barcode_image = code128.image(item_code)
                image_stream = io.BytesIO()
                barcode_image.save(image_stream, format='PNG')
                image_stream.seek(0)

                # Insert barcode image into the Excel file
                img = OpenpyxlImage(image_stream)
                cell_address = f'C{ws.max_row}'  # Calculate the cell address where the image will be placed
                img.anchor = cell_address  # Anchor the image using the cell address as a string
                ws.add_image(img)

                # Set the row height to accommodate the barcode image
                ws.row_dimensions[ws.max_row].height = 90  # Adjust height as needed

        # Save the workbook
        wb.save(excel_file_path)
        messagebox.showinfo("Success", f"Barcode information saved to {excel_file_path}.")

    def scan_barcode(self):
        os.system('py scan.py') 

if __name__ == "__main__":
    root = tk.Tk()
    app = BarcodeApp(root)
    root.mainloop()
